"""`crawlbtc trace <address>` - follow the money out from an address.

Builds a directed address-flow graph outward from a starting address to a
bounded depth, with per-edge value estimated by the haircut model (each
recipient's share weighted by the sender's contribution to that
transaction). Reports the origin's incoming funders and outgoing
recipients, detects flows that loop back to the origin, and writes:

  <addr>_trace.html   self-contained interactive graph (send-able, offline)
  <addr>_trace.xlsx   nodes / edges / origin in / origin out / summary
  <addr>_trace.json   the raw graph

Everything is read-only and uses indexed lookups. High-degree "hub"
addresses (exchanges etc.) are detected and NOT expanded, so the graph
cannot explode. All caps are reported so nothing is silently dropped.
"""

import datetime
import json
import os
import sys
from collections import deque

import psycopg

SATS = 100_000_000

# An address with more than this many outputs is treated as a hub
# (exchange/service/mixer) and not expanded further.
HUB_UTXO_THRESHOLD = 20_000


def _connect(cfg):
    return psycopg.connect(cfg.db_conninfo, autocommit=True)


def _utxo_count(cur, address):
    cur.execute("""
        SELECT COUNT(*) FROM blockchain.transaction_io
        WHERE address = %s AND io_type = 'out';
    """, (address,))
    return cur.fetchone()[0]


def _outgoing_edges(cur, address, fanout):
    """Top `fanout` recipients of value sent FROM `address`, haircut-weighted."""
    cur.execute("""
        WITH u_utxos AS (
            SELECT txid, idx, amount
              FROM blockchain.transaction_io
             WHERE address = %s AND io_type = 'out'
        ),
        spent AS (
            SELECT s.spending_txid, SUM(u.amount) AS u_in
              FROM u_utxos u
              JOIN blockchain.spends s
                ON s.prev_txid = u.txid AND s.prev_vout = u.idx
             GROUP BY s.spending_txid
        ),
        tx_out AS (
            SELECT sp.spending_txid, o.address AS to_addr, SUM(o.amount) AS v_out
              FROM spent sp
              JOIN blockchain.transaction_io o
                ON o.txid = sp.spending_txid AND o.io_type = 'out'
             WHERE o.address IS NOT NULL
             GROUP BY sp.spending_txid, o.address
        )
        SELECT t.to_addr,
               SUM(t.v_out * (sp.u_in::numeric / NULLIF(tx.total_in, 0)))::bigint AS est_value,
               COUNT(DISTINCT t.spending_txid) AS tx_count,
               MIN(tx.received_time) AS first_seen,
               MAX(tx.received_time) AS last_seen
          FROM tx_out t
          JOIN spent sp ON sp.spending_txid = t.spending_txid
          JOIN blockchain.transactions tx ON tx.txid = t.spending_txid
         GROUP BY t.to_addr
         ORDER BY est_value DESC NULLS LAST
         LIMIT %s;
    """, (address, fanout))
    return cur.fetchall()


def _incoming_edges(cur, address, fanout):
    """Top `fanout` funders that paid value INTO `address`."""
    cur.execute("""
        WITH u_recv AS (
            SELECT DISTINCT txid
              FROM blockchain.transaction_io
             WHERE address = %s AND io_type = 'out'
        )
        SELECT i.address AS from_addr,
               SUM(i.amount)::bigint AS value_in,
               COUNT(DISTINCT i.txid) AS tx_count,
               MIN(tx.received_time) AS first_seen,
               MAX(tx.received_time) AS last_seen
          FROM blockchain.transaction_io i
          JOIN u_recv r ON r.txid = i.txid
          JOIN blockchain.transactions tx ON tx.txid = i.txid
         WHERE i.io_type = 'in' AND i.address IS NOT NULL
         GROUP BY i.address
         ORDER BY value_in DESC NULLS LAST
         LIMIT %s;
    """, (address, fanout))
    return cur.fetchall()


def _entity_lookup(cur, addresses):
    """Return {address: (entity_name, category, source)} for known entities.

    Prefers sanctioned > mixer > everything else when an address has tags
    from multiple sources. Returns {} if the entity_tags table is absent.
    """
    if not addresses:
        return {}
    try:
        cur.execute("""
            SELECT DISTINCT ON (address) address, entity_name, category, source
            FROM blockchain.entity_tags
            WHERE address = ANY(%s)
            ORDER BY address,
                     CASE category WHEN 'sanctioned' THEN 0 WHEN 'mixer' THEN 1
                                   WHEN 'exchange' THEN 2 ELSE 3 END,
                     confidence DESC NULLS LAST;
        """, (list(addresses),))
        return {a: (name, cat, src) for a, name, cat, src in cur.fetchall()}
    except psycopg.errors.UndefinedTable:
        return {}


def _addr_totals(cur, address):
    """Total received / sent (sats) and utxo count for an address."""
    cur.execute("""
        SELECT COALESCE(SUM(amount) FILTER (WHERE io_type='out'), 0),
               COALESCE(SUM(amount) FILTER (WHERE io_type='in'), 0)
          FROM blockchain.transaction_io
         WHERE address = %s;
    """, (address,))
    recv, sent = cur.fetchone()
    return int(recv), int(sent)


def _origin_cluster(cur, origin, tx_cap=3000, addr_cap=8000):
    """Addresses probably owned by the SAME entity as `origin`.

    Common-input-ownership heuristic: addresses co-spent as inputs in the
    same transaction are almost certainly one wallet. One bounded round
    (direct co-inputs) - transitive expansion is intentionally NOT done, as
    it over-merges through CoinJoin/PayJoin. Always includes the origin.
    Bounded by tx_cap/addr_cap and the session statement_timeout.
    """
    try:
        cur.execute("""
            WITH origin_spends AS (
              SELECT DISTINCT s.spending_txid
                FROM blockchain.transaction_io o
                JOIN blockchain.spends s ON s.prev_txid = o.txid AND s.prev_vout = o.idx
               WHERE o.address = %s AND o.io_type = 'out'
               LIMIT %s
            )
            SELECT DISTINCT i.address
              FROM origin_spends os
              JOIN blockchain.transaction_io i
                ON i.txid = os.spending_txid AND i.io_type = 'in'
             WHERE i.address IS NOT NULL
             LIMIT %s;
        """, (origin, tx_cap, addr_cap))
        return {r[0] for r in cur.fetchall()} | {origin}
    except psycopg.errors.QueryCanceled:
        return {origin}


def build_graph(cfg, origin, depth, fanout, max_nodes, direction="out", cluster=True):
    with _connect(cfg) as conn:
        cur = conn.cursor()
        cur.execute("SET statement_timeout = '120s';")

        nodes = {}
        edges = []
        loops = []   # edges that return value to the origin's cluster (round-trips)

        cluster_addrs = _origin_cluster(cur, origin) if cluster else {origin}

        def ensure_node(addr, d, side):
            if addr not in nodes:
                recv, sent = _addr_totals(cur, addr)
                nodes[addr] = {
                    "address": addr, "depth": abs(d), "side": side,
                    "level": 0 if addr == origin else (abs(d) if side == "out" else -abs(d)),
                    "is_origin": addr == origin,
                    "total_received_btc": recv / SATS,
                    "total_sent_btc": sent / SATS,
                    "same_owner": (addr in cluster_addrs and addr != origin),
                    "is_hub": False, "truncated": False,
                }
            return nodes[addr]

        ensure_node(origin, 0, "origin")

        def expand(dir_):
            q = deque([(origin, 0)])
            seen = set()
            while q:
                addr, d = q.popleft()
                if addr in seen or d >= depth:
                    continue
                seen.add(addr)
                if _utxo_count(cur, addr) > HUB_UTXO_THRESHOLD and addr != origin:
                    nodes[addr]["is_hub"] = True
                    continue
                rows = (_outgoing_edges if dir_ == "out" else _incoming_edges)(cur, addr, fanout)
                if len(rows) >= fanout:
                    nodes[addr]["truncated"] = True
                for other, est_value, tx_count, first_seen, last_seen in rows:
                    if other is None:
                        continue
                    ensure_node(other, d + 1, dir_)
                    frm, to = (addr, other) if dir_ == "out" else (other, addr)
                    returns_to_owner = (other in cluster_addrs)
                    edge = {
                        "from": frm, "to": to,
                        "value_btc": (est_value or 0) / SATS,
                        "tx_count": tx_count,
                        "first_seen": str(first_seen) if first_seen else None,
                        "last_seen": str(last_seen) if last_seen else None,
                        "dir": dir_,
                        "confidence": round(max(0.15, 0.8 ** d), 3),
                        "returns_to_owner": returns_to_owner,
                    }
                    edges.append(edge)
                    if returns_to_owner:
                        loops.append(edge)
                    if len(nodes) >= max_nodes:
                        q.clear()
                        break
                    if d + 1 < depth and other not in cluster_addrs:
                        q.append((other, d + 1))

        if direction in ("out", "both"):
            expand("out")
        if direction in ("in", "both"):
            expand("in")

        # Origin's immediate funders, for the incoming-context panel.
        origin_incoming = [{
            "address": fr, "value_btc": (v or 0) / SATS, "tx_count": tc,
            "first_seen": str(fs) if fs else None, "last_seen": str(ls) if ls else None,
            "same_owner": fr in cluster_addrs,
        } for fr, v, tc, fs, ls in _incoming_edges(cur, origin, fanout)]

        # Known-entity labels.
        for addr, (name, cat, src) in _entity_lookup(cur, nodes.keys()).items():
            nodes[addr]["entity"] = name
            nodes[addr]["entity_category"] = cat
            nodes[addr]["entity_source"] = src

        o = nodes[origin]
        related = [n for n in nodes.values() if n["same_owner"]]
        return {
            "origin": origin,
            "generated": datetime.datetime.now().isoformat(),
            "params": {"depth": depth, "fanout": fanout, "max_nodes": max_nodes,
                       "direction": direction, "clustering": cluster,
                       "hub_threshold": HUB_UTXO_THRESHOLD},
            "origin_summary": {
                "total_received_btc": o["total_received_btc"],
                "total_sent_btc": o["total_sent_btc"],
                "distinct_funders_shown": len(origin_incoming),
                "related_wallets": len(related),
                "cluster_size": len(cluster_addrs),
                "round_trips_to_owner": len(loops),
                "node_count": len(nodes),
                "edge_count": len(edges),
                "capped": len(nodes) >= max_nodes,
            },
            "nodes": list(nodes.values()),
            "edges": edges,
            "origin_incoming": origin_incoming,
            "related_wallets": [n["address"] for n in related],
            "loops": loops,
        }


# ---------- outputs ----------

def write_json(graph, path):
    with open(path, "w") as f:
        json.dump(graph, f, indent=2)


def write_xlsx(graph, path):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    bold = Font(bold=True)

    ws = wb.active
    ws.title = "Summary"
    s = graph["origin_summary"]
    rows = [
        ("crawlbtc address trace", ""),
        ("origin", graph["origin"]),
        ("generated", graph["generated"]),
        ("direction", graph["params"]["direction"]),
        ("depth", graph["params"]["depth"]),
        ("max fan-out per node", graph["params"]["fanout"]),
        ("", ""),
        ("total received (BTC)", round(s["total_received_btc"], 8)),
        ("total sent (BTC)", round(s["total_sent_btc"], 8)),
        ("funders shown", s["distinct_funders_shown"]),
        ("addresses in graph", s["node_count"]),
        ("flows (edges)", s["edge_count"]),
        ("related wallets (same owner)", s["related_wallets"]),
        ("origin cluster size", s["cluster_size"]),
        ("round-trips to owner", s["round_trips_to_owner"]),
        ("graph hit node cap", s["capped"]),
    ]
    for r in rows:
        ws.append(r)
    ws["A1"].font = bold

    ws = wb.create_sheet("Nodes")
    ws.append(["address", "level", "side", "is_origin", "same_owner", "entity",
               "category", "is_hub", "total_received_btc", "total_sent_btc"])
    for c in ws[1]:
        c.font = bold
    for n in sorted(graph["nodes"], key=lambda x: (x["level"], -x["total_received_btc"])):
        ws.append([n["address"], n["level"], n["side"], n["is_origin"], n["same_owner"],
                   n.get("entity", ""), n.get("entity_category", ""), n["is_hub"],
                   round(n["total_received_btc"], 8), round(n["total_sent_btc"], 8)])

    ws = wb.create_sheet("Flows (edges)")
    ws.append(["from", "to", "dir", "est_value_btc", "tx_count", "confidence",
               "returns_to_owner", "first_seen", "last_seen"])
    for c in ws[1]:
        c.font = bold
    for e in sorted(graph["edges"], key=lambda x: -x["value_btc"]):
        ws.append([e["from"], e["to"], e["dir"], round(e["value_btc"], 8), e["tx_count"],
                   e["confidence"], e["returns_to_owner"], e["first_seen"], e["last_seen"]])

    ws = wb.create_sheet("Origin incoming")
    ws.append(["funder_address", "value_btc", "tx_count", "same_owner", "first_seen", "last_seen"])
    for c in ws[1]:
        c.font = bold
    for f in graph["origin_incoming"]:
        ws.append([f["address"], round(f["value_btc"], 8), f["tx_count"],
                   f.get("same_owner", False), f["first_seen"], f["last_seen"]])

    if graph["related_wallets"]:
        ws = wb.create_sheet("Related wallets")
        ws.append(["address (probably same owner as origin)"])
        ws["A1"].font = bold
        for a in graph["related_wallets"]:
            ws.append([a])

    wb.save(path)


def write_html(graph, path):
    data_json = json.dumps({
        "origin": graph["origin"],
        "nodes": [{"id": n["address"], "depth": n["depth"], "level": n["level"],
                   "origin": n["is_origin"], "hub": n["is_hub"], "owner": n["same_owner"],
                   "entity": n.get("entity"), "etype": n.get("entity_category"),
                   "recv": round(n["total_received_btc"], 6),
                   "sent": round(n["total_sent_btc"], 6)} for n in graph["nodes"]],
        "edges": [{"s": e["from"], "t": e["to"], "v": round(e["value_btc"], 6),
                   "n": e["tx_count"], "loop": e["returns_to_owner"]} for e in graph["edges"]],
    })
    summary = graph["origin_summary"]
    html = _HTML_TEMPLATE.replace("__DATA__", data_json) \
        .replace("__ORIGIN__", graph["origin"]) \
        .replace("__GENERATED__", graph["generated"]) \
        .replace("__DEPTH__", str(graph["params"]["depth"])) \
        .replace("__NODES__", str(summary["node_count"])) \
        .replace("__EDGES__", str(summary["edge_count"])) \
        .replace("__LOOPS__", str(summary["round_trips_to_owner"])) \
        .replace("__RECV__", f"{summary['total_received_btc']:.8f}") \
        .replace("__SENT__", f"{summary['total_sent_btc']:.8f}")
    with open(path, "w") as f:
        f.write(html)


def cmd_trace(args, cfg):
    origin = args.address.strip()
    out_dir = os.path.abspath(os.path.expanduser(args.out or os.getcwd()))
    os.makedirs(out_dir, exist_ok=True)

    # Warn about combinatorial explosion: worst case ~ fanout^depth nodes.
    worst = args.fanout ** args.depth
    if worst > args.max_nodes:
        print(f"note: depth {args.depth} x fanout {args.fanout} could reach ~{worst:,} "
              f"addresses but is capped at {args.max_nodes:,}. For deep traces lower "
              f"--fanout (e.g. --fanout 3 --depth {args.depth} = {3**args.depth:,}) so the "
              f"picture isn't truncated. High-degree hubs are pruned automatically.")

    print(f"tracing {origin} (direction={args.direction}, depth={args.depth}, "
          f"fanout={args.fanout}) ...")
    try:
        graph = build_graph(cfg, origin, args.depth, args.fanout, args.max_nodes,
                            direction=args.direction, cluster=not args.no_cluster)
    except psycopg.errors.QueryCanceled:
        print("a query timed out (statement_timeout 120s). Try smaller --fanout/--depth, "
              "or run when build-balances/backup are not competing for disk.", file=sys.stderr)
        sys.exit(1)

    s = graph["origin_summary"]
    print(f"  received {s['total_received_btc']:.8f} BTC, sent {s['total_sent_btc']:.8f} BTC")
    print(f"  graph: {s['node_count']} addresses, {s['edge_count']} flows")
    print(f"  related wallets (same owner): {s['related_wallets']} "
          f"(origin cluster {s['cluster_size']}), round-trips to owner: {s['round_trips_to_owner']}"
          + ("  [hit node cap]" if s["capped"] else ""))

    base = os.path.join(out_dir, f"{origin}_trace")
    write_json(graph, base + ".json")
    write_html(graph, base + ".html")
    try:
        write_xlsx(graph, base + ".xlsx")
        xlsx_note = base + ".xlsx"
    except ImportError:
        xlsx_note = "(openpyxl not installed - skipped .xlsx; pip install openpyxl)"

    print("\nwrote:")
    print(f"  {base}.html   (open / send this)")
    print(f"  {xlsx_note}")
    print(f"  {base}.json")


_HTML_TEMPLATE = r"""<!doctype html>
<html><head><meta charset="utf-8"><title>crawlbtc trace __ORIGIN__</title>
<style>
  :root{color-scheme:light dark}
  body{margin:0;font:13px/1.4 system-ui,sans-serif;background:#0e1116;color:#e6edf3}
  header{padding:10px 14px;background:#161b22;border-bottom:1px solid #30363d}
  header b{color:#58a6ff}
  .stats{display:flex;gap:18px;flex-wrap:wrap;margin-top:6px;color:#9da7b3}
  .stats span{white-space:nowrap}
  #wrap{display:flex;height:calc(100vh - 62px)}
  svg{flex:1;background:#0e1116}
  #side{width:320px;padding:12px;background:#161b22;border-left:1px solid #30363d;overflow:auto}
  #side h3{margin:0 0 6px;font-size:13px;color:#58a6ff}
  #side .a{word-break:break-all;font-family:ui-monospace,monospace;font-size:12px}
  .k{color:#9da7b3}.v{color:#e6edf3}
  line.edge{stroke:#3d4551;stroke-opacity:.6}
  line.loop{stroke:#f85149;stroke-width:2;stroke-opacity:.9}
  circle{cursor:pointer;stroke:#0e1116;stroke-width:1.5}
  text.lbl{fill:#8b949e;font-size:9px;pointer-events:none}
  .legend span{margin-right:12px}
  .dot{display:inline-block;width:10px;height:10px;border-radius:50%;vertical-align:middle;margin-right:4px}
</style></head>
<body>
<header>
  <div>crawlbtc trace — origin <b>__ORIGIN__</b></div>
  <div class="stats">
    <span>received <b>__RECV__</b> BTC</span>
    <span>sent <b>__SENT__</b> BTC</span>
    <span>depth __DEPTH__</span>
    <span>__NODES__ addresses</span>
    <span>__EDGES__ flows</span>
    <span>loops to origin: <b>__LOOPS__</b></span>
    <span class="legend"><span class="dot" style="background:#f0b72f"></span>origin
      <span class="dot" style="background:#58a6ff"></span>depth1
      <span class="dot" style="background:#3fb950"></span>depth2
      <span class="dot" style="background:#a371f7"></span>depth3+
      <span class="dot" style="background:#db6d28"></span>hub</span>
    <span style="color:#6e7681">generated __GENERATED__</span>
  </div>
</header>
<div id="wrap"><svg id="g"></svg>
  <div id="side"><h3>click a node</h3><div id="info" class="k">Nodes are addresses; edges follow value outward. Red edges loop back to the origin. Drag to move, scroll to zoom.</div></div>
</div>
<script>
const DATA = __DATA__;
const svg = document.getElementById('g');
const side = document.getElementById('info');
const W = () => svg.clientWidth, H = () => svg.clientHeight;
const COLORS = ['#f0b72f','#58a6ff','#3fb950','#a371f7'];
function color(n){ if(n.hub) return '#db6d28'; if(n.origin) return COLORS[0]; return COLORS[Math.min(n.depth,3)]; }
function radius(n){ return n.origin?12:Math.max(4,6+Math.log10((n.recv||0)+1)); }

const nodes = DATA.nodes.map(n=>({...n,x:W()/2+(Math.random()-.5)*300,y:H()/2+(Math.random()-.5)*300,vx:0,vy:0}));
const idx = {}; nodes.forEach(n=>idx[n.id]=n);
const links = DATA.edges.filter(e=>idx[e.s]&&idx[e.t]).map(e=>({...e,source:idx[e.s],target:idx[e.t]}));

const NS='http://www.w3.org/2000/svg';
let vb={x:0,y:0,w:0,h:0,k:1};
function el(t,a){const e=document.createElementNS(NS,t);for(const k in a)e.setAttribute(k,a[k]);return e;}
const gEdges=el('g'),gNodes=el('g'),gLabels=el('g');
svg.append(gEdges,gNodes,gLabels);
const lineEls=links.map(l=>{const e=el('line',{class:l.loop?'loop':'edge'});gEdges.append(e);return e;});
const circleEls=nodes.map(n=>{const c=el('circle',{r:radius(n),fill:color(n)});c.addEventListener('click',()=>showInfo(n));c.addEventListener('mousedown',ev=>startDrag(ev,n));gNodes.append(c);return c;});
const labelEls=nodes.map(n=>{const t=el('text',{class:'lbl'});t.textContent=(n.origin?'★ ':'')+n.id.slice(0,8)+'…';gLabels.append(t);return t;});

function showInfo(n){
  side.innerHTML='<div class="a">'+n.id+'</div><br>'+
    row('depth',n.depth)+row('received (BTC)',n.recv)+row('sent (BTC)',n.sent)+
    (n.origin?'<br><b style="color:#f0b72f">ORIGIN</b>':'')+(n.hub?'<br><b style="color:#db6d28">HUB (not expanded)</b>':'');
  circleEls.forEach((c,i)=>c.setAttribute('stroke',nodes[i]===n?'#fff':'#0e1116'));
}
function row(k,v){return '<div><span class="k">'+k+':</span> <span class="v">'+v+'</span></div>';}

// force sim
function tick(){
  for(const n of nodes){ if(n===drag) continue;
    for(const m of nodes){ if(n===m)continue; let dx=n.x-m.x,dy=n.y-m.y,d2=dx*dx+dy*dy+.01; if(d2<40000){const f=800/d2;n.vx+=dx*f/Math.sqrt(d2);n.vy+=dy*f/Math.sqrt(d2);} }
  }
  for(const l of links){ let dx=l.target.x-l.source.x,dy=l.target.y-l.source.y,d=Math.sqrt(dx*dx+dy*dy)+.01,f=(d-90)*0.01;
    if(l.source!==drag){l.source.vx+=dx/d*f;l.source.vy+=dy/d*f;} if(l.target!==drag){l.target.vx-=dx/d*f;l.target.vy-=dy/d*f;} }
  for(const n of nodes){ if(n===drag)continue; n.vx+=(W()/2-n.x)*0.001;n.vy+=(H()/2-n.y)*0.001; n.x+=n.vx*=0.85;n.y+=n.vy*=0.85; }
  links.forEach((l,i)=>{lineEls[i].setAttribute('x1',l.source.x);lineEls[i].setAttribute('y1',l.source.y);lineEls[i].setAttribute('x2',l.target.x);lineEls[i].setAttribute('y2',l.target.y);});
  nodes.forEach((n,i)=>{circleEls[i].setAttribute('cx',n.x);circleEls[i].setAttribute('cy',n.y);labelEls[i].setAttribute('x',n.x+radius(n)+2);labelEls[i].setAttribute('y',n.y+3);});
  requestAnimationFrame(tick);
}
// drag + zoom
let drag=null,dox,doy;
function startDrag(ev,n){drag=n;const p=pt(ev);dox=p.x-n.x;doy=p.y-n.y;ev.preventDefault();}
window.addEventListener('mousemove',ev=>{if(drag){const p=pt(ev);drag.x=p.x-dox;drag.y=p.y-doy;drag.vx=drag.vy=0;}});
window.addEventListener('mouseup',()=>drag=null);
function pt(ev){const r=svg.getBoundingClientRect();return {x:(ev.clientX-r.left-vb.x)/vb.k,y:(ev.clientY-r.top-vb.y)/vb.k};}
let pan=false,px,py;
svg.addEventListener('mousedown',ev=>{if(ev.target===svg){pan=true;px=ev.clientX-vb.x;py=ev.clientY-vb.y;}});
window.addEventListener('mousemove',ev=>{if(pan){vb.x=ev.clientX-px;vb.y=ev.clientY-py;apply();}});
window.addEventListener('mouseup',()=>pan=false);
svg.addEventListener('wheel',ev=>{ev.preventDefault();const s=ev.deltaY<0?1.1:0.9;vb.k*=s;apply();});
function apply(){gEdges.setAttribute('transform',`translate(${vb.x},${vb.y}) scale(${vb.k})`);gNodes.setAttribute('transform',gEdges.getAttribute('transform'));gLabels.setAttribute('transform',gEdges.getAttribute('transform'));}
tick();
</script>
</body></html>
"""
